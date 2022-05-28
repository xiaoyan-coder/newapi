import openpyxl


class HandleExcel:

    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def read_excel(self):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        rows = list(sheet.rows)
        header = [i.value for i in rows[0]]
        list_data = []
        for j in rows[1:]:
            data1 = [m.value for m in j]
            data = dict(zip(header, data1))
            list_data.append(data)
        return list_data

    def write_excel(self, row, column, value):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        sheet.cell(row=row, column=column, value=value)
        wb.save(self.file_path)


if __name__ == '__main__':
    file_name = r'E:\newapi\datas\read_excel.xlsx'
    a = Handle_Excel(file_name, 'login')
    a.write_excel(1,1,'case_id')

